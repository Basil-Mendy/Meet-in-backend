import random
import re
import string
from datetime import datetime
from difflib import SequenceMatcher
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.utils.html import strip_tags
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import VerificationRequest
from forums.email_service import EmailService
from forums.models import Forum, ForumMembership as ForumForumMembership
from forums.serializers import ForumSerializer as CommunityForumSerializer
from .models import School, SchoolForum, SchoolMembership, IndependentForumRequest, AdminRole, SchoolJoinRequest, SchoolForumMessage, ForumRequest
from .serializers import (
    SchoolSerializer,
    SchoolForumSerializer,
    SchoolMembershipSerializer,
    SchoolForumMemberSerializer,
    IndependentForumRequestSerializer,
    AdminRoleSerializer,
    UserSummarySerializer,
    SchoolJoinRequestSerializer,
    SchoolForumMessageSerializer,
    ForumRequestSerializer,
    SchoolForumAboutSerializer,
)
import csv
import io
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def _is_admin_user(user):
    admin_role = getattr(user, "admin_role", None)
    return bool(admin_role or user.is_staff or user.is_superuser)


def _approved_members_count(forum):
    return SchoolMembership.objects.filter(forum=forum, status="APPROVED").count()


def _is_forum_member(user, forum):
    return SchoolMembership.objects.filter(user=user, forum=forum, status="APPROVED").exists()


def _is_forum_moderator(user, forum):
    return SchoolMembership.objects.filter(
        user=user,
        forum=forum,
        status="APPROVED",
        role__in=("MODERATOR", "PATRON", "CHAIRMAN", "SECRETARY"),
    ).exists()


def _is_school_owner(user, forum):
    try:
        return forum.school.created_by_id == user.id
    except Exception:
        return False


def _can_review_join_request(user, forum):
    if _is_admin_user(user):
        return True
    if _is_school_owner(user, forum):
        return True
    if _approved_members_count(forum) < 10:
        return _is_forum_member(user, forum)
    return _is_forum_moderator(user, forum)


class SchoolViewSet(viewsets.ModelViewSet):
    queryset = School.objects.all().order_by("name")
    serializer_class = SchoolSerializer
    permission_classes = [IsAuthenticatedOrReadOnly]

    def get_queryset(self):
        queryset = School.objects.all().order_by("name")

        q = self.request.query_params.get("q", "").strip()
        state = self.request.query_params.get("state", "").strip()
        lga = self.request.query_params.get("lga", "").strip()
        ward = self.request.query_params.get("ward", "").strip()
        country = self.request.query_params.get("country", "").strip()

        if q:
            queryset = queryset.filter(
                Q(name__icontains=q)
                | Q(address__icontains=q)
                | Q(description__icontains=q)
                | Q(country__icontains=q)
                | Q(state__icontains=q)
                | Q(lga__icontains=q)
                | Q(ward__icontains=q)
                | Q(main_contact_number__icontains=q)
            ).distinct()
        if state:
            queryset = queryset.filter(state__icontains=state)
        if lga:
            queryset = queryset.filter(lga__icontains=lga)
        if ward:
            queryset = queryset.filter(ward__icontains=ward)
        if country:
            queryset = queryset.filter(country__icontains=country)

        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(created_by=request.user, is_approved=True, is_verified=True)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SchoolForumViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SchoolForum.objects.all().order_by("year", "name")
    serializer_class = SchoolForumSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=["get"], url_path="members")
    def members(self, request, *args, **kwargs):
        forum = self.get_object()
        memberships = SchoolMembership.objects.filter(forum=forum, status="APPROVED").select_related("user", "school")
        serializer = SchoolForumMemberSerializer(memberships, many=True, context={"request": request})
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path=r"members/(?P<member_id>[^/.]+)/assign-role")
    def assign_role(self, request, *args, **kwargs):
        forum = self.get_object()
        member_id = kwargs.get("member_id")
        membership = get_object_or_404(SchoolMembership, id=member_id, forum=forum)
        role = (request.data.get("role") or "MEMBER").upper()

        valid_roles = [choice[0] for choice in SchoolMembership.ROLE_CHOICES]
        if role not in valid_roles:
            return Response({"error": "Invalid role."}, status=status.HTTP_400_BAD_REQUEST)

        requester = SchoolMembership.objects.filter(
            user=request.user,
            forum=forum,
            status="APPROVED",
        ).first()
        if not requester or requester.role == "MEMBER":
            return Response({"error": "You do not have permission to assign roles."}, status=status.HTTP_403_FORBIDDEN)

        membership.role = role
        membership.save(update_fields=["role"])
        return Response(SchoolForumMemberSerializer(membership, context={"request": request}).data)

    @action(detail=True, methods=["delete"], url_path=r"members/(?P<member_id>[^/.]+)")
    def remove_member(self, request, *args, **kwargs):
        forum = self.get_object()
        member_id = kwargs.get("member_id")
        membership = get_object_or_404(SchoolMembership, id=member_id, forum=forum)

        requester = SchoolMembership.objects.filter(
            user=request.user,
            forum=forum,
            status="APPROVED",
        ).first()
        if not requester or requester.role == "MEMBER":
            return Response({"error": "You do not have permission to remove members."}, status=status.HTTP_403_FORBIDDEN)

        membership.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class JoinSchoolForumView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, forum_id):
        forum = SchoolForum.objects.get(id=forum_id)
        # Expect graduation_year and certificate (optional) in multipart data
        graduation_year = request.data.get("graduation_year")
        certificate = request.FILES.get("certificate")

        # Prevent duplicate requests if still pending or already approved
        existing = SchoolJoinRequest.objects.filter(user=request.user, forum=forum).order_by("-requested_at").first()
        if existing and existing.status in ("PENDING", "APPROVED"):
            payload = {"created": False, "status": existing.status}
            return Response(payload, status=status.HTTP_200_OK)

        # Allow new request if previous was rejected or none exists
        join_req = SchoolJoinRequest.objects.create(
            user=request.user,
            forum=forum,
            graduation_year=graduation_year or None,
            certificate=certificate,
        )
        approved_count = _approved_members_count(forum)
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            admins = User.objects.filter(Q(is_superuser=True) | Q(admin_role__is_super_admin=True)).distinct()
            if approved_count < 10:
                reviewers = User.objects.filter(school_memberships__forum=forum, school_memberships__status="APPROVED").distinct()
                recipients = list(admins) + list(reviewers)
            else:
                reviewers = User.objects.filter(
                    school_memberships__forum=forum,
                    school_memberships__status="APPROVED",
                    school_memberships__role__in=("MODERATOR", "PATRON", "CHAIRMAN", "SECRETARY"),
                ).distinct()
                recipients = list(admins) + list(reviewers)

            emails = []
            seen = set()
            for user in recipients:
                if user.email and user.email not in seen:
                    emails.append(user.email)
                    seen.add(user.email)

            if emails:
                subject = f"New join request for {forum.name}"
                html = f"<p>{request.user.get_full_name() or request.user.email} requested to join {forum.name}.</p><p>Please review the request in the admin dashboard.</p>"
                EmailService.send_system_notification(subject=subject, html_body=html, text_body=strip_tags(html), to=emails)
        except Exception:
            pass
        serializer = SchoolJoinRequestSerializer(join_req, context={"request": request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class JoinMateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        forum_id = request.data.get("forum_id")
        role = (request.data.get("role") or "MEMBER").upper()
        forum = SchoolForum.objects.get(id=forum_id)

        initial_status = "APPROVED" if role in {"CHAIRMAN", "SECRETARY"} else "PENDING"

        membership, created = SchoolMembership.objects.get_or_create(
            user=request.user,
            forum=forum,
            defaults={"school": forum.school, "status": initial_status, "role": role},
        )
        if not created:
            membership.status = initial_status
            membership.role = role
            membership.save(update_fields=["status", "role"])

        joined_general_forum = False
        if role in {"CHAIRMAN", "SECRETARY"}:
            general_forum = SchoolForum.objects.filter(school=forum.school, is_general=True).first()
            if general_forum:
                general_membership, general_created = SchoolMembership.objects.get_or_create(
                    user=request.user,
                    forum=general_forum,
                    defaults={"school": forum.school, "status": "APPROVED", "role": role},
                )
                if not general_created:
                    general_membership.status = "APPROVED"
                    general_membership.role = role
                    general_membership.save(update_fields=["status", "role"])
                joined_general_forum = True

        return Response({
            "created": created,
            "status": membership.status,
            "role": membership.role,
            "joined_general_forum": joined_general_forum,
        }, status=status.HTTP_200_OK)


class MyJoinRequestsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        records = SchoolJoinRequest.objects.filter(user=request.user).order_by("-requested_at")
        serializer = SchoolJoinRequestSerializer(records, many=True, context={"request": request})
        return Response(serializer.data)


class MyOwnedSchoolsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        schools = School.objects.filter(created_by=request.user).prefetch_related("forums")
        payload = []
        for school in schools:
            forums_payload = []
            for forum in school.forums.all():
                member_count = SchoolMembership.objects.filter(forum=forum, status="APPROVED").count()
                pending_count = SchoolJoinRequest.objects.filter(forum=forum, status="PENDING").count()
                forums_payload.append({
                    "id": str(forum.id),
                    "name": forum.name,
                    "year": forum.year,
                    "is_general": forum.is_general,
                    "description": forum.description,
                    "member_count": member_count,
                    "pending_requests": pending_count,
                })
            payload.append({
                "id": str(school.id),
                "name": school.name,
                "main_contact_number": school.main_contact_number,
                "description": school.description,
                "forum_count": school.forums.count(),
                "forums": forums_payload,
            })
        return Response(payload)


class MyOwnedSchoolJoinRequestsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        school_ids = School.objects.filter(created_by=request.user).values_list("id", flat=True)
        records = SchoolJoinRequest.objects.filter(forum__school_id__in=school_ids, status="PENDING").select_related("forum", "user").order_by("-requested_at")
        serializer = SchoolJoinRequestSerializer(records, many=True, context={"request": request})
        return Response(serializer.data)


class MySchoolForumsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        memberships = SchoolMembership.objects.filter(user=request.user, status="APPROVED").select_related("forum", "forum__school")
        payload = []
        for membership in memberships:
            payload.append({
                "id": str(membership.forum.id),
                "name": membership.forum.name,
                "description": membership.forum.description,
                "is_general": membership.forum.is_general,
                "year": membership.forum.year,
                "school_id": str(membership.forum.school.id),
                "school_name": membership.forum.school.name,
                "role": membership.role,
                "type": "school",
            })
        return Response(payload)


class SchoolForumMessageView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, forum_id):
        try:
            forum = SchoolForum.objects.get(id=forum_id)
        except SchoolForum.DoesNotExist:
            return Response({"detail": "Forum not found."}, status=status.HTTP_404_NOT_FOUND)

        messages = SchoolForumMessage.objects.filter(forum=forum).select_related("sender").order_by("-created_at")
        serializer = SchoolForumMessageSerializer(messages, many=True, context={"request": request})
        return Response(serializer.data)

    def post(self, request, forum_id):
        try:
            forum = SchoolForum.objects.get(id=forum_id)
        except SchoolForum.DoesNotExist:
            return Response({"detail": "Forum not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = SchoolForumMessageSerializer(data=request.data, context={"request": request})
        if serializer.is_valid():
            message = serializer.save(forum=forum, sender=request.user)
            return Response(SchoolForumMessageSerializer(message, context={"request": request}).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SchoolForumAboutView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, forum_id):
        forum = get_object_or_404(SchoolForum, id=forum_id)
        is_member = SchoolMembership.objects.filter(forum=forum, user=request.user, status="APPROVED").exists()
        is_school_owner = forum.school.created_by_id == request.user.id

        if not (is_member or is_school_owner):
            return Response({"detail": "You are not a member of this forum"}, status=status.HTTP_403_FORBIDDEN)

        serializer = SchoolForumAboutSerializer(forum, context={"request": request})
        return Response(serializer.data)


class SchoolForumMyRoleView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, forum_id):
        forum = get_object_or_404(SchoolForum, id=forum_id)
        membership = SchoolMembership.objects.filter(forum=forum, user=request.user, status="APPROVED").first()
        if not membership:
            return Response({"error": "You are not a member of this forum"}, status=status.HTTP_404_NOT_FOUND)

        return Response({"role": membership.role})


class ForumRequestView(APIView):
    permission_classes = [IsAuthenticated]

    def _normalize_text(self, value):
        return re.sub(r"\W+", " ", (value or "").lower()).strip()

    def _similarity_score(self, left, right):
        left_text = self._normalize_text(left)
        right_text = self._normalize_text(right)
        if not left_text or not right_text:
            return 0.0
        return SequenceMatcher(None, left_text, right_text).ratio()

    def _find_similar_organizations(self, payload):
        organization_name = (payload.get("organization_name") or "").strip()
        address = (payload.get("address") or "").strip()
        phone = (payload.get("contact_phone") or "").strip()
        state = (payload.get("state") or "").strip()
        country = (payload.get("country") or "").strip()

        candidates = Forum.objects.all()
        query = Q()
        if organization_name:
            name_tokens = [token for token in re.split(r"\W+", organization_name.lower()) if len(token) > 2]
            for token in name_tokens:
                query |= Q(name__icontains=token)
        if address:
            query |= Q(address__icontains=address)
        if phone:
            query |= Q(phone__iexact=phone) | Q(contact_phone__iexact=phone)

        if query:
            candidates = candidates.filter(query)

        if state:
            candidates = candidates.filter(state__icontains=state)
        if country:
            candidates = candidates.filter(country__icontains=country)

        suggestions = []
        for forum in candidates.distinct():
            name_similarity = self._similarity_score(organization_name, forum.name)
            address_similarity = self._similarity_score(address, forum.address)
            phone_match = phone and (phone == forum.phone or phone == forum.contact_phone)

            if phone_match or name_similarity >= 0.88 or (name_similarity >= 0.78 and address_similarity >= 0.60):
                score = max(name_similarity, address_similarity)
                suggestions.append((score, forum))

        suggestions.sort(key=lambda item: item[0], reverse=True)
        return [forum for _, forum in suggestions[:10]]

    def _validate_duplicate_request(self, payload):
        request_type = (payload.get("request_type") or "SCHOOL").upper()
        school_name = (payload.get("school_name") or "").strip()
        organization_name = (payload.get("organization_name") or "").strip()

        if request_type == "SCHOOL" and school_name:
            existing_school = School.objects.filter(name__iexact=school_name).first()
            if existing_school:
                return Response({"detail": "This school already exists."}, status=status.HTTP_400_BAD_REQUEST)

            pending_request = ForumRequest.objects.filter(
                request_type="SCHOOL",
                school_name__iexact=school_name,
                status__in=["PENDING", "UNDER_REVIEW", "APPROVED"],
            ).first()
            if pending_request:
                return Response({"detail": "A request for this forum is already under review."}, status=status.HTTP_400_BAD_REQUEST)

        if request_type == "ORGANIZATION" and organization_name:
            existing_orgs = Forum.objects.filter(name__iexact=organization_name)
            if existing_orgs.exists():
                existing_forum = existing_orgs.first()
                exact_matches = list(existing_orgs)
                similar_candidates = self._find_similar_organizations(payload)
                exact_ids = {forum.id for forum in exact_matches}
                suggestions = exact_matches + [forum for forum in similar_candidates if forum.id not in exact_ids]
                payloads = [
                    {
                        "id": str(f.id),
                        "name": f.name,
                        "description": f.description,
                        "address": f.address,
                        "state": f.state,
                        "lga": f.lga,
                        "country": f.country,
                        "phone": f.phone,
                        "contact_person": f.contact_person,
                        "contact_phone": f.contact_phone,
                        "contact_email": f.contact_email,
                        "join_policy": f.join_policy,
                        "created_by_name": getattr(f.created_by, "first_name", "") or getattr(f.created_by, "username", "") or "Unknown",
                    }
                    for f in suggestions
                ]
                return Response({
                    "detail": "This organization already exists.",
                    "existing_forum": {
                        "id": str(existing_forum.id),
                        "name": existing_forum.name,
                        "description": existing_forum.description,
                        "address": existing_forum.address,
                        "state": existing_forum.state,
                        "lga": existing_forum.lga,
                        "country": existing_forum.country,
                        "phone": existing_forum.phone,
                        "contact_person": existing_forum.contact_person,
                        "contact_phone": existing_forum.contact_phone,
                        "contact_email": existing_forum.contact_email,
                        "join_policy": existing_forum.join_policy,
                        "created_by_name": getattr(existing_forum.created_by, "first_name", "") or getattr(existing_forum.created_by, "username", "") or "Unknown",
                    },
                    "existing_forums": [
                        {
                            "id": str(f.id),
                            "name": f.name,
                            "description": f.description,
                            "address": f.address,
                            "state": f.state,
                            "lga": f.lga,
                            "country": f.country,
                            "phone": f.phone,
                            "contact_person": f.contact_person,
                            "contact_phone": f.contact_phone,
                            "contact_email": f.contact_email,
                            "join_policy": f.join_policy,
                            "created_by_name": getattr(f.created_by, "first_name", "") or getattr(f.created_by, "username", "") or "Unknown",
                        }
                        for f in exact_matches
                    ],
                    "similar_forums": payloads,
                }, status=status.HTTP_409_CONFLICT)

            pending_request = ForumRequest.objects.filter(
                request_type="ORGANIZATION",
                organization_name__iexact=organization_name,
                status__in=["PENDING", "UNDER_REVIEW", "APPROVED"],
            ).first()
            if pending_request:
                return Response({"detail": "A request for this forum is already under review."}, status=status.HTTP_400_BAD_REQUEST)

            similar_orgs = self._find_similar_organizations(payload)
            if similar_orgs:
                suggestions = [
                    {
                        "id": str(f.id),
                        "name": f.name,
                        "description": f.description,
                        "address": f.address,
                        "state": f.state,
                        "lga": f.lga,
                        "country": f.country,
                        "phone": f.phone,
                        "contact_person": f.contact_person,
                        "contact_phone": f.contact_phone,
                        "contact_email": f.contact_email,
                        "join_policy": f.join_policy,
                        "created_by_name": getattr(f.created_by, "first_name", "") or getattr(f.created_by, "username", "") or "Unknown",
                    }
                    for f in similar_orgs
                ]
                return Response(
                    {
                        "detail": "A similar organization already exists.",
                        "similar_forums": suggestions,
                    },
                    status=status.HTTP_409_CONFLICT,
                )

        return None

    def post(self, request):
        payload = request.data.copy()
        payload["submitted_by"] = request.user.id
        payload["status"] = "PENDING"

        duplicate_response = self._validate_duplicate_request(payload)
        if duplicate_response is not None:
            return duplicate_response

        serializer = ForumRequestSerializer(data=payload)
        if serializer.is_valid():
            obj = serializer.save(submitted_by=request.user)

            if obj.request_type == "ORGANIZATION":
                organization_name = obj.organization_name or obj.school_name
                forum_id = "".join(random.choices(string.ascii_uppercase + string.digits, k=12))
                while Forum.objects.filter(forum_id=forum_id).exists():
                    forum_id = "".join(random.choices(string.ascii_uppercase + string.digits, k=12))

                organization_forum = Forum.objects.create(
                    forum_id=forum_id,
                    name=organization_name,
                    description=f"Organization forum for {organization_name}",
                    address=obj.address,
                    state=obj.state,
                    lga=obj.lga,
                    country=obj.country,
                    email=obj.contact_email,
                    phone=obj.contact_phone,
                    contact_person=obj.contact_person,
                    contact_phone=obj.contact_phone,
                    contact_email=obj.contact_email,
                    join_policy=obj.join_policy or "CLOSED",
                    is_completed=True,
                    is_verified=True,
                    is_searchable=True,
                    created_by=request.user,
                )

                ForumForumMembership.objects.create(
                    user=request.user,
                    forum=organization_forum,
                    role="P",
                    is_active=True,
                )

                # Mark request approved immediately
                obj.status = "APPROVED"
                obj.reviewed_by = request.user
                obj.reviewed_at = timezone.now()
                obj.save(update_fields=["status", "reviewed_by", "reviewed_at"])

                return Response({"detail": "Community forum created successfully.", "forum": {
                    "id": str(organization_forum.id),
                    "forum_id": organization_forum.forum_id,
                    "name": organization_forum.name,
                    "description": organization_forum.description,
                    "address": organization_forum.address,
                    "contact_person": organization_forum.contact_person,
                    "contact_phone": organization_forum.contact_phone,
                    "contact_email": organization_forum.contact_email,
                }}, status=status.HTTP_201_CREATED)

            return Response(ForumRequestSerializer(obj).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        records = ForumRequest.objects.order_by("-created_at")
        return Response(ForumRequestSerializer(records, many=True).data)


class ForumRequestDecisionView(APIView):
    permission_classes = [IsAuthenticated]

    def _school_forum_payload(self, school, year):
        if year is None:
            return {
                "school": school,
                "year": None,
                "is_general": True,
                "name": f"{school.name} General Alumni Forum",
                "description": f"General alumni forum for {school.name}",
            }
        return {
            "school": school,
            "year": year,
            "is_general": False,
            "name": f"Class of {year}",
            "description": f"Alumni forum for {school.name} class of {year}",
        }

    def post(self, request, request_id):
        if not _is_admin_user(request.user):
            return Response({"detail": "You are not an admin."}, status=status.HTTP_403_FORBIDDEN)

        record = ForumRequest.objects.get(id=request_id)
        action = request.data.get("action", "approve")
        record.remarks = (request.data.get("remarks") or "").strip() or record.remarks

        if action == "approve":
            if record.request_type == "SCHOOL":
                if not record.school_name:
                    return Response({"detail": "School name is required to approve this request."}, status=status.HTTP_400_BAD_REQUEST)
                try:
                    School.objects.get(name__iexact=record.school_name)
                except School.DoesNotExist:
                    pass
                else:
                    return Response({"detail": "A school with this name already exists."}, status=status.HTTP_400_BAD_REQUEST)

                with transaction.atomic():
                    school = School.objects.create(
                        name=record.school_name,
                        address=record.address,
                        country=record.country or "Nigeria",
                        state=record.state,
                        lga=record.lga,
                        year_established=record.year_established or timezone.now().year,
                        school_type=record.school_type or "SECONDARY",
                        created_by=record.submitted_by,
                        is_approved=True,
                        is_verified=True,
                    )
                    school.ensure_default_forums()

                    for forum in school.forums.all():
                        SchoolMembership.objects.get_or_create(
                            user=record.submitted_by,
                            forum=forum,
                            defaults={"school": school, "status": "APPROVED", "role": "MODERATOR"},
                        )

                    record.status = "APPROVED"
                    record.reviewed_by = request.user
                    record.reviewed_at = timezone.now()
                    record.save(update_fields=["status", "reviewed_by", "reviewed_at", "remarks"])

                try:
                    if record.submitted_by and record.submitted_by.email:
                        EmailService.send_system_notification(
                            subject="Congratulations! Your request has been approved.",
                            html_body="<p>Congratulations!</p><p>Your request has been approved.</p><p>The forum is now available for members to join.</p>",
                            text_body="Congratulations!\nYour request has been approved.\nThe forum is now available for members to join.",
                            to=[record.submitted_by.email],
                        )
                except Exception:
                    pass

            else:
                organization_name = record.organization_name or record.school_name
                forum_id = "".join(random.choices(string.ascii_uppercase + string.digits, k=12))
                while Forum.objects.filter(forum_id=forum_id).exists():
                    forum_id = "".join(random.choices(string.ascii_uppercase + string.digits, k=12))

                with transaction.atomic():
                    organization_forum = Forum.objects.create(
                        forum_id=forum_id,
                        name=organization_name,
                        description=f"Organization forum for {organization_name}",
                        address=record.address,
                        email=record.contact_email,
                        phone=record.phone,
                        is_completed=True,
                        is_verified=True,
                        is_searchable=True,
                        created_by=record.submitted_by,
                    )
                    ForumForumMembership.objects.create(
                        user=record.submitted_by,
                        forum=organization_forum,
                        role="P",
                        is_active=True,
                    )

                    record.status = "APPROVED"
                    record.reviewed_by = request.user
                    record.reviewed_at = timezone.now()
                    record.save(update_fields=["status", "reviewed_by", "reviewed_at", "remarks"])

                try:
                    if record.submitted_by and record.submitted_by.email:
                        EmailService.send_system_notification(
                            subject="Congratulations! Your request has been approved.",
                            html_body="<p>Congratulations!</p><p>Your request has been approved.</p><p>The forum is now available for members to join.</p>",
                            text_body="Congratulations!\nYour request has been approved.\nThe forum is now available for members to join.",
                            to=[record.submitted_by.email],
                        )
                except Exception:
                    pass

        elif action == "reject":
            record.status = "REJECTED"
            record.reviewed_by = request.user
            record.reviewed_at = timezone.now()
            record.save(update_fields=["status", "reviewed_by", "reviewed_at", "remarks"])
            try:
                if record.submitted_by and record.submitted_by.email:
                    EmailService.send_system_notification(
                        subject="Unfortunately your request could not be approved.",
                        html_body=f"<p>Unfortunately your request could not be approved.</p><p>Reason: {record.remarks or 'No reason provided.'}</p>",
                        text_body=f"Unfortunately your request could not be approved.\nReason: {record.remarks or 'No reason provided.'}",
                        to=[record.submitted_by.email],
                    )
            except Exception:
                pass

        return Response(ForumRequestSerializer(record).data)


class IndependentForumRequestView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = IndependentForumRequestSerializer(data=request.data)
        if serializer.is_valid():
            obj = serializer.save(user=request.user, requested_by=request.user)
            return Response(IndependentForumRequestSerializer(obj).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        records = IndependentForumRequest.objects.order_by("-created_at")
        return Response(IndependentForumRequestSerializer(records, many=True).data)


class SchoolApprovalView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, school_id):
        school = School.objects.get(id=school_id)
        action = request.data.get("action", "approve")
        requested_role = (request.data.get("role") or "").upper()
        if action == "approve":
            school.is_approved = True
            school.is_verified = True
        else:
            school.is_approved = False
            school.is_verified = False
        school.save()
        return Response(SchoolSerializer(school).data)


class IndependentForumDecisionView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, request_id):
        record = IndependentForumRequest.objects.get(id=request_id)
        action = request.data.get("action", "approve")
        record.status = "APPROVED" if action == "approve" else "REJECTED"
        record.reviewed_at = datetime.now()
        record.save()
        return Response(IndependentForumRequestSerializer(record).data)


class VerificationRequestListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _is_admin_user(request.user):
            return Response({"detail": "You are not an admin."}, status=status.HTTP_403_FORBIDDEN)

        records = VerificationRequest.objects.select_related("user").order_by("-submitted_at")
        payload = []
        for record in records:
            payload.append({
                "id": record.id,
                "status": record.status,
                "submitted_at": record.submitted_at,
                "reviewed_at": record.reviewed_at,
                "id_type": record.id_type,
                "user_id": record.user.id,
                "user_name": f"{record.user.first_name} {record.user.last_name}".strip(),
                "user_email": record.user.email,
                "is_user_verified": record.user.is_verified,
            })
        return Response(payload)


class VerificationRequestDecisionView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, request_id):
        if not _is_admin_user(request.user):
            return Response({"detail": "You do not have permission to review verification requests."}, status=status.HTTP_403_FORBIDDEN)

        verification_request = VerificationRequest.objects.get(id=request_id)
        action = request.data.get("action", "approve")
        verification_request.status = "approved" if action == "approve" else "rejected"
        verification_request.reviewed_at = datetime.now()
        verification_request.reviewed_by = request.user
        verification_request.save(update_fields=["status", "reviewed_at", "reviewed_by"])

        verification_request.user.is_verified = action == "approve"
        verification_request.user.save(update_fields=["is_verified"])

        return Response({
            "id": verification_request.id,
            "status": verification_request.status,
            "reviewed_at": verification_request.reviewed_at,
            "submitted_at": verification_request.submitted_at,
            "user_verified": verification_request.user.is_verified,
        })


class AdminUserListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _is_admin_user(request.user):
            return Response({"detail": "You are not an admin."}, status=status.HTTP_403_FORBIDDEN)

        queryset = get_user_model().objects.all().order_by("first_name", "last_name", "email")
        q = request.query_params.get("q", "").strip()
        verified = request.query_params.get("verified", "").strip().lower()

        if q:
            queryset = queryset.filter(
                Q(first_name__icontains=q)
                | Q(last_name__icontains=q)
                | Q(email__icontains=q)
                | Q(phone__icontains=q)
            )

        if verified in ["true", "false"]:
            queryset = queryset.filter(is_verified=(verified == "true"))

        users = queryset
        return Response(UserSummarySerializer(users, many=True).data)


class AdminUserVerificationView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, user_id):
        if not _is_admin_user(request.user):
            return Response({"detail": "You are not an admin."}, status=status.HTTP_403_FORBIDDEN)

        try:
            user = get_user_model().objects.get(id=user_id)
        except get_user_model().DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        action = request.data.get("action", "verify")

        if action == "verify":
            user.is_verified = True
        elif action == "unverify":
            user.is_verified = False
        else:
            return Response({"detail": "Action must be verify or unverify."}, status=status.HTTP_400_BAD_REQUEST)

        user.save(update_fields=["is_verified"])
        return Response(UserSummarySerializer(user).data)


class AdminCommunityForumsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not _is_admin_user(request.user):
            return Response({"detail": "You are not an admin."}, status=status.HTTP_403_FORBIDDEN)

        queryset = Forum.objects.all().select_related("created_by").order_by("-created_at")
        verified = request.query_params.get("verified", "").strip().lower()
        q = request.query_params.get("q", "").strip()

        if verified in ["true", "false"]:
            queryset = queryset.filter(is_verified=(verified == "true"))
        if q:
            queryset = queryset.filter(Q(name__icontains=q) | Q(description__icontains=q) | Q(address__icontains=q) | Q(state__icontains=q) | Q(lga__icontains=q))

        payload = []
        for forum in queryset:
            member_count = ForumForumMembership.objects.filter(forum=forum, is_active=True).count()
            payload.append({
                "id": str(forum.id),
                "name": forum.name,
                "description": forum.description,
                "address": forum.address,
                "state": forum.state,
                "lga": forum.lga,
                "country": forum.country,
                "contact_person": forum.contact_person,
                "contact_phone": forum.contact_phone,
                "contact_email": forum.contact_email,
                "join_policy": forum.join_policy,
                "is_verified": forum.is_verified,
                "is_completed": forum.is_completed,
                "is_searchable": forum.is_searchable,
                "member_count": member_count,
                "created_at": forum.created_at,
                "created_by": getattr(forum.created_by, "email", None),
            })
        return Response(payload)


class AdminDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        admin_role = getattr(request.user, "admin_role", None)
        if not _is_admin_user(request.user):
            return Response({"detail": "You are not an admin."}, status=status.HTTP_403_FORBIDDEN)

        admin_role = admin_role or type("AdminRoleFallback", (), {"is_super_admin": request.user.is_superuser, "can_manage_schools": True, "can_verify_users": True, "can_manage_forums": True})()
        pending_schools = School.objects.filter(is_approved=False).count() + ForumRequest.objects.filter(request_type="SCHOOL", status__in=["PENDING", "UNDER_REVIEW"]).count()
        pending_forums = IndependentForumRequest.objects.filter(status="PENDING").count()
        pending_verifications = VerificationRequest.objects.filter(status="pending").count()
        return Response({
            "is_super_admin": admin_role.is_super_admin,
            "can_manage_schools": admin_role.can_manage_schools,
            "can_verify_users": admin_role.can_verify_users,
            "can_manage_forums": admin_role.can_manage_forums,
            "pending_schools": pending_schools,
            "pending_forums": pending_forums,
            "pending_verifications": pending_verifications,
            "today": datetime.now().date().isoformat(),
        })


class AdminManagementView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Only the super admin can manage other admins."}, status=status.HTTP_403_FORBIDDEN)

        admins = []
        for user in get_user_model().objects.filter(is_staff=True).order_by("first_name", "last_name", "email"):
            admin_role = getattr(user, "admin_role", None)
            admins.append({
                "id": str(user.id),
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "phone": user.phone,
                "is_super_admin": bool(user.is_superuser or (admin_role and admin_role.is_super_admin)),
                "is_active": user.is_active,
            })
        return Response(admins)

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Only the super admin can create other admins."}, status=status.HTTP_403_FORBIDDEN)

        email = request.data.get("email")
        password = request.data.get("password")
        if not email or not password:
            return Response({"detail": "Email and password are required."}, status=status.HTTP_400_BAD_REQUEST)

        user = get_user_model().objects.create_user(
            email=email,
            password=password,
            phone=request.data.get("phone", ""),
            first_name=request.data.get("first_name", "Admin"),
            last_name=request.data.get("last_name", "User"),
            is_staff=True,
            is_superuser=False,
            is_verified=True,
        )
        admin_role = AdminRole.objects.create(
            user=user,
            is_super_admin=request.data.get("is_super_admin", False),
            can_manage_schools=True,
            can_verify_users=True,
            can_manage_forums=True,
        )
        return Response({
            "id": str(user.id),
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "phone": user.phone,
            "is_super_admin": admin_role.is_super_admin,
        }, status=status.HTTP_201_CREATED)

    def delete(self, request, admin_id=None):
        if not request.user.is_superuser:
            return Response({"detail": "Only the super admin can remove other admins."}, status=status.HTTP_403_FORBIDDEN)

        if not admin_id:
            return Response({"detail": "Admin id is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            admin_user = get_user_model().objects.get(id=admin_id)
        except get_user_model().DoesNotExist:
            return Response({"detail": "Admin not found."}, status=status.HTTP_404_NOT_FOUND)

        if admin_user.is_superuser or admin_user == request.user:
            return Response({"detail": "You cannot delete the super admin or yourself."}, status=status.HTTP_400_BAD_REQUEST)

        admin_role = getattr(admin_user, "admin_role", None)
        if admin_role:
            admin_role.delete()
        admin_user.is_staff = False
        admin_user.is_active = False
        admin_user.save(update_fields=["is_staff", "is_active"])
        return Response({"detail": "Admin removed successfully."}, status=status.HTTP_200_OK)

class AdminSchoolManagementView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Only super admins or staff with can_manage_schools can create directly
        admin_role = getattr(request.user, "admin_role", None)
        if not (request.user.is_superuser or (admin_role and admin_role.can_manage_schools)):
            return Response({"detail": "You are not authorized to create schools."}, status=status.HTTP_403_FORBIDDEN)

        data = request.data
        name = (data.get("name") or "").strip()
        if not name:
            return Response({"detail": "Name is required."}, status=status.HTTP_400_BAD_REQUEST)

        # prevent duplicates
        if School.objects.filter(name__iexact=name).exists():
            return Response({"detail": "School already exists."}, status=status.HTTP_400_BAD_REQUEST)

        school = School.objects.create(
            name=name,
            address=data.get("address", ""),
            country=data.get("country", "Nigeria"),
            state=data.get("state", ""),
            lga=data.get("lga", ""),
            ward=data.get("ward", ""),
            year_established=data.get("year_established") or None,
            school_type=data.get("school_type") or "SECONDARY",
            primary_color=data.get("primary_color") or "#0f172a",
            secondary_color=data.get("secondary_color") or "#f59e0b",
            description=data.get("description", ""),
            main_contact_number=data.get("main_contact_number", ""),
            created_by=request.user,
            is_approved=True,
            is_verified=True,
            visibility=data.get("visibility") or "PUBLIC",
        )
        school.ensure_default_forums()
        return Response(SchoolSerializer(school).data, status=status.HTTP_201_CREATED)


class AdminSchoolBulkUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        admin_role = getattr(request.user, "admin_role", None)
        if not (request.user.is_superuser or (admin_role and admin_role.can_manage_schools)):
            return Response({"detail": "You are not authorized to upload schools."}, status=status.HTTP_403_FORBIDDEN)

        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        filename = upload.name.lower()
        created = 0
        skipped = 0
        errors = []

        try:
            if filename.endswith(".csv"):
                text = io.TextIOWrapper(upload.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif filename.endswith(('.xlsx', '.xls')) and OPENPYXL_AVAILABLE:
                wb = openpyxl.load_workbook(upload, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: row[i] for i in range(len(headers))})
            else:
                return Response({"detail": "Unsupported file type or missing openpyxl for Excel files."}, status=status.HTTP_400_BAD_REQUEST)

            for idx, row in enumerate(rows, start=1):
                try:
                    name = (row.get('name') or row.get('school_name') or '').strip()
                    if not name:
                        skipped += 1
                        continue
                    if School.objects.filter(name__iexact=name).exists():
                        skipped += 1
                        continue

                    year = row.get('year_established') or row.get('year') or None
                    try:
                        year = int(year) if year else None
                    except Exception:
                        year = None

                    visibility = (row.get('visibility') or 'PUBLIC').upper()
                    if visibility not in dict(School.VISIBILITY_CHOICES):
                        visibility = 'PUBLIC'

                    school = School.objects.create(
                        name=name,
                        address=row.get('address') or '',
                        country=row.get('country') or 'Nigeria',
                        state=row.get('state') or '',
                        lga=row.get('lga') or '',
                        ward=row.get('ward') or '',
                        year_established=year or None,
                        school_type=(row.get('school_type') or 'SECONDARY'),
                        primary_color=row.get('primary_color') or '#0f172a',
                        secondary_color=row.get('secondary_color') or '#f59e0b',
                        description=row.get('description') or '',
                        main_contact_number=row.get('main_contact_number') or '',
                        created_by=request.user,
                        is_approved=True,
                        is_verified=True,
                        visibility=visibility,
                    )
                    school.ensure_default_forums()
                    created += 1
                except Exception as e:
                    errors.append({"row": idx, "error": str(e)})

        except Exception as e:
            return Response({"detail": "Failed to parse file", "error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"created": created, "skipped": skipped, "errors": errors})


class AdminRoleViewSet(viewsets.ModelViewSet):
    queryset = AdminRole.objects.all()
    serializer_class = AdminRoleSerializer
    permission_classes = [IsAuthenticated]


def _is_forum_patron(user, forum):
    try:
        membership = SchoolMembership.objects.filter(user=user, forum=forum, status="APPROVED", role="PATRON").first()
        return bool(membership)
    except Exception:
        return False


def _is_forum_executive(user, forum):
    try:
        membership = SchoolMembership.objects.filter(
            user=user,
            forum=forum,
            status="APPROVED",
            role__in=("PATRON", "CHAIRMAN", "SECRETARY", "MODERATOR"),
        ).first()
        return bool(membership)
    except Exception:
        return False


class AdminSchoolJoinRequestsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if _is_admin_user(request.user):
            records = list(SchoolJoinRequest.objects.select_related("forum", "user").order_by("-requested_at"))
        else:
            records = []
            for record in SchoolJoinRequest.objects.select_related("forum", "user").order_by("-requested_at"):
                if _can_review_join_request(request.user, record.forum):
                    records.append(record)

        first_time = request.query_params.get("first_time")
        if first_time and first_time.lower() in ("1", "true", "yes"):
            records = [r for r in records if not SchoolMembership.objects.filter(forum=r.forum, status="APPROVED").exists()]

        serializer = SchoolJoinRequestSerializer(records, many=True, context={"request": request})
        return Response(serializer.data)


class AdminSchoolJoinRequestDecisionView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, request_id):
        try:
            jr = SchoolJoinRequest.objects.get(id=request_id)
        except SchoolJoinRequest.DoesNotExist:
            return Response({"detail": "Request not found."}, status=status.HTTP_404_NOT_FOUND)

        # allow reviewer to request a specific role for the approve action
        requested_role = (request.data.get("role") or "").upper()

        # only admins or approved forum reviewers can approve/reject
        can_review = _can_review_join_request(request.user, jr.forum)
        if not can_review:
            return Response({"detail": "Not authorized to review this request."}, status=status.HTTP_403_FORBIDDEN)

        action = request.data.get("action", "approve")
        if action == "approve":
            approved_count = _approved_members_count(jr.forum)
            role = "MEMBER"
            if requested_role:
                role = requested_role
            elif approved_count < 10:
                role = "MODERATOR"

            membership, created = SchoolMembership.objects.get_or_create(
                user=jr.user,
                forum=jr.forum,
                defaults={"school": jr.forum.school, "status": "APPROVED", "role": role},
            )
            if not created:
                membership.status = "APPROVED"
                membership.role = role
                membership.save(update_fields=["status", "role"])

            # If role is CHAIRMAN or SECRETARY ensure membership in the school's general forum as well
            try:
                if role in {"CHAIRMAN", "SECRETARY"}:
                    general_forum = SchoolForum.objects.filter(school=jr.forum.school, is_general=True).first()
                    if general_forum:
                        gen_membership, gen_created = SchoolMembership.objects.get_or_create(
                            user=jr.user,
                            forum=general_forum,
                            defaults={"school": general_forum.school, "status": "APPROVED", "role": role},
                        )
                        if not gen_created:
                            gen_membership.status = "APPROVED"
                            gen_membership.role = role
                            gen_membership.save(update_fields=["status", "role"])
            except Exception:
                pass

            jr.status = "APPROVED"
            jr.rejection_reason = None
            message = "approved"
        elif action == "reject":
            jr.status = "REJECTED"
            # save rejection reason if provided
            rejection_reason = (request.data.get("reason") or "").strip() or None
            if rejection_reason:
                jr.rejection_reason = rejection_reason
            message = "rejected"
        else:
            return Response({"detail": "Action must be approve or reject."}, status=status.HTTP_400_BAD_REQUEST)

        jr.reviewed_by = request.user
        jr.reviewed_at = timezone.now()
        jr.save()

        try:
            # notify requester by email, include rejection reason when present
            if jr.user and jr.user.email:
                subj = f"Your join request for {jr.forum.name} has been {jr.status.lower()}"
                html = f"<p>Your request to join {jr.forum.name} was {jr.status.lower()} by {request.user.get_full_name() or request.user.email}.</p>"
                if jr.status == "REJECTED" and jr.rejection_reason:
                    html += f"<p>Reason: {jr.rejection_reason}</p>"
                EmailService.send_system_notification(subject=subj, html_body=html, text_body=strip_tags(html), to=[jr.user.email])
        except Exception:
            pass

        return Response({"detail": f"Request {message}.", "status": jr.status})
